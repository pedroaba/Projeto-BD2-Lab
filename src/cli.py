import collections
import os
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill, Alignment

from src.controller.points_controller import Controller


class SimpleCLI:
    def __init__(self):
        self.commands = {}

    def add_command(self, name, function):
        self.commands[name] = function

    def run(self):
        output_command_message = None
        while True:
            for command_index, command in enumerate(
                list(self.commands.keys()) + ['quit']
            ):
                print(f"{command_index + 1} - {command}")
            print()

            command = input("Enter a command: ")
            if command == "quit":
                print("Goodbye!")
                exit(0)
            elif command in self.commands:
                output_command_message = self.commands[command]()
                if output_command_message is not None:
                    separator = "-" * 50

                    print(separator)
                    print(f'Saída do comando anterior: {output_command_message}', end=f'\n{separator}')
                    output_command_message = None
            else:
                print("Invalid command. Try again.")


class PointerCLI(SimpleCLI):
    def __init__(self):
        super().__init__()
        self.controller = Controller()

        self.add_command('register_employee', self.register_employee)
        self.add_command('register_point', self.register_point)
        self.add_command('resume', self.resume)
        self.add_command('update_employee_info', self.update_employee_info)

        self.run()

    def register_employee(self):
        cpf = input("Entre com o CPF: ")
        email = input("Entre com o E-mail: ")
        name = input("Entre com o Nome: ")

        try:
            self.controller.add_employee(cpf, email, name)
            return f"Empregado com o CPF: {cpf} foi cadastrado com sucesso"
        except Exception as e:
            return f"Não foi possível cadastrar o empregado {name}"

    def register_point(self):
        cpf = input("Entre com o CPF: ")
        try:
            self.controller.add_points(cpf)
            return "Ponto Registrado com sucesso"
        except Exception as e:
            return "Não foi possível registrar o ponto!"

    def resume(self):
        try:
            points = self.controller.get_employees_points()
        except Exception as e:
            return "Não foi possível buscar os registros de ponto!"

        if isinstance(points, collections.abc.Iterable):
            filename = Path('.') / 'Resumo de pontos dos funcionários.xlsx'

            points_df = pd.DataFrame(data=points)
            points_df.to_excel(filename, index=False, sheet_name='Registros')

            wb = load_workbook(filename)
            ws = wb.get_sheet_by_name('Registros')

            cell_header_bg = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for colindex, coltext in zip(range(5), ['MongoDB ID', 'ID', 'CPF do Empregado', 'Início', 'Término']):
                cell = ws.cell(1, colindex + 1)
                cell.value = coltext
                cell.fill = cell_header_bg

            for letter_col in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[letter_col].width = 40
                ws.column_dimensions[letter_col].alignment = Alignment(horizontal="center", vertical="center")
            for row_number in range(1, points_df.shape[0], 1):
                ws.row_dimensions[row_number].height = 16
                ws.row_dimensions[row_number].alignment = Alignment(horizontal="center", vertical="center")

            wb.save(filename)
            print(f"Resumo exportado em '{filename.absolute()}'")
        else:
            print(points)

    def update_employee_info(self):
        cpf = input("Entre com o CPF: ")

        try:
            employee = self.controller.get_employee_by_cpf(cpf)
            name = input("Entre com o nome do empregado: ")

            try:
                self.controller.update_employee_name(employee['cpf'], name)
            except Exception as e:
                return f"Não foi possível encontrar o empregado: '{employee['name']}'!"
        except Exception as e:
            return f"Empregado com o cpf: '{cpf}' não foi encontrado!"

    def clear_employee_points(self):
        cpf = input("Entre com o CPF: ")

        try:
            self.controller.remove_points(cpf)
        except Exception as e:
            return f"Empregado com o cpf: '{cpf}' não foi encontrado!"
